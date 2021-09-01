"""
Tem de ser executado no WLS
"""

import concurrent.futures
import datetime
import ipaddress
import logging
import os
import subprocess
import sys
import time
import re

import netmiko
import openpyxl
from netmiko.ssh_exception import (AuthenticationException,
                                   NetMikoTimeoutException)
from paramiko.ssh_exception import SSHException

# ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ Logging ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬
# criar um logger específico deste módulo
main_logger = logging.getLogger(__name__)
main_logger.setLevel(logging.DEBUG)  # definir o nível de verbosidade do logger

# criar o formato dos logs a aplicar no file_formatter
file_formatter = logging.Formatter(
    '%(asctime)s:%(levelname)s:%(name)s:%(message)s')
# criar o formato dos logs a aplicar no stream_handler
steam_formatter = logging.Formatter('%(message)s')

file_handler = logging.FileHandler(
    f'{os.path.join("Logs",os.path.splitext(os.path.basename(__file__))[0])}.log')  # criar ficheiro de logs
# definir o nível de verbosidade do file_handler
file_handler.setLevel(logging.DEBUG)
# aplicar o formato de log anteriormente criado
file_handler.setFormatter(file_formatter)

# criar stream_handler que serve para mostrar logs noutros canais
stream_handler = logging.StreamHandler()
# definir o nível de verbosidade do stream_handler
stream_handler.setLevel(logging.INFO)
# aplicar o formato de log anteriormente criado
stream_handler.setFormatter(steam_formatter)

main_logger.addHandler(file_handler)  # aplicar file_handler ao logger
main_logger.addHandler(stream_handler)  # aplicar stream_handler ao logger
# ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬ End Logging ▬▬▬▬▬▬▬▬▬▬▬▬▬▬▬


def validate_ip(ip):
    try:
        ipaddress.ip_address(ip)
    except:
        return False
    else:
        return True


def cls():
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["clear"])
    elif sys.platform == "win32":
        subprocess.run(["cls"], shell=True)


def createTempFile(pathToExcelFile, cwd, excelFileName):
    if sys.platform in ("linux", "darwin"):
        subprocess.run(["cp", pathToExcelFile, cwd, excelFileName])
    elif sys.platform == "win32":
        subprocess.run(["robocopy", pathToExcelFile, cwd,
                       excelFileName], shell=True, stdout=subprocess.DEVNULL)


def write_output(equipment):
    with open(f'{os.path.join("Outputs",os.path.splitext(os.path.basename(__file__))[0])}.txt', 'at') as f:
        f.write(f"{equipment['secret']}\t({equipment['ip']})\n")


def write_error(equipment):
    with open(f'{os.path.join("Outputs",os.path.splitext(os.path.basename(__file__))[0])}-ERROR.txt', 'at') as f:
        f.write(f"{equipment['secret']}\t({equipment['ip']})\n")
        # main_logger.info(f"Error: {equipment['secret']} ({equipment['ip']})")


def main():
    # clear output file
    with open(f'{os.path.join("Outputs",os.path.splitext(os.path.basename(__file__))[0])}.txt', 'w') as f:
        pass
    # clear output file 2
    with open(f'{os.path.join("Outputs",os.path.splitext(os.path.basename(__file__))[0])}-ERROR.txt', 'w') as f:
        pass

    global cfg_file
    if sys.platform in ("linux", "darwin"):
        excel = "/mnt/c/Users/joao.caldeira.ext/OneDrive - Portugal Telecom/Trabalho/COS/SPMS/Scripting/cadastro-caldeira.xlsm"
        cfg_file = 'Config files/Routers-RIS_remover_LAN-tecnico.ios'
    elif sys.platform == "win32":
        excel = r"C:\Users\joao.caldeira.ext\OneDrive - Portugal Telecom\Trabalho\COS\SPMS\Scripting\cadastro-caldeira.xlsm"
        cfg_file = r'Config files\Routers-RIS_SNMP_correction.ios'

    excelFileName = os.path.basename(excel)
    pathToExcelFile = os.path.dirname(excel)

    username = 'jcaldeira'
    password = os.getenv('PWD_TACACS_RIS_ALTICE')

    try:
        wb = openpyxl.load_workbook(excel, data_only=True, read_only=True)
        delTempFileFlag = False
    except PermissionError as e:
        print(f"File is open, so cannot be accessed:\n{e.filename}\n")
        createTempFile(pathToExcelFile, os.getcwd(), excelFileName)
        print("Temporary copy of file created to work with\n")
        wb = openpyxl.load_workbook(
            excelFileName, data_only=True, read_only=True)
        delTempFileFlag = True
    except FileNotFoundError as e:
        sys.exit(f"File not found: {e.filename}")

    sheet = wb["Parsed-Altice"]
    device_list = []
    models = {
        'ASR1001': 'Cisco',
        'ASR1001-x c/10G': 'Cisco',
        'BUN-ACESS-TELDAT-4GE+ROUTING': 'Teldat',
        'C1101-4P': 'Cisco',
        'C1111-8P': 'Cisco',
        'ISR4351': 'Cisco',
        'ISR4451': 'Cisco',
        'ISR-4451X': 'Cisco',
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        site_id = str(row[0])
        ip_pri = str(row[9])
        ip_bck = str(row[10])
        model_pri = str(row[5])
        model_bck = str(row[6])

        # validar IP e confirmar se é para testar ("SIM")
        if validate_ip(ip_pri) and str(row[12]).upper() == 'SIM':
            if models.get(f'{model_pri}') == 'Cisco':
                equipment = {
                    'device_type': 'cisco_xe',
                    'ip': ip_pri,
                    'username': username,
                    'password': password,
                    'secret': site_id,
                    'ssh_config_file': '~/.ssh/config',
                    # 'global_delay_factor': 2,
                }
                equipment['session_log'] = f"""{os.path.join('Logs','Session Logs',f"{equipment['secret']} ({ip_pri})")}.log"""
                device_list.append(equipment)

        if validate_ip(ip_bck) and str(row[13]).upper() == 'SIM':
            if models.get(f'{model_bck}') == 'Cisco':
                equipment = {
                    'device_type': 'cisco_xe',
                    'ip': ip_bck,
                    'username': username,
                    'password': password,
                    'secret': site_id,
                    'ssh_config_file': '~/.ssh/config',
                    # 'global_delay_factor': 2,
                }
                equipment['session_log'] = f"""{os.path.join('Logs','Session Logs',f"{equipment['secret']} ({ip_bck})")}.log"""
                device_list.append(equipment)

    main_logger.debug(f'{device_list = }')

    print(f'Equipments to work with: {len(device_list)}')
    input(f'Config file is correct: {cfg_file}\nPress <ENTER> to continue...')

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
        executor.map(connect_and_commands, device_list)

    wb.close()
    if delTempFileFlag:
        os.remove(excelFileName)

    return len(device_list)


def connect_and_commands(equipment):
    pattern_to_search = r"(10.255.\d{1,3}.\d{1,3}/30)"
    compiled_pattern_to_search = re.compile(pattern_to_search)

    main_logger.info(f"Accessing: {equipment['secret']} ({equipment['ip']})")
    try:
        with netmiko.ConnectHandler(
                **equipment,
                conn_timeout=60,
                auth_timeout=60,
                banner_timeout=60,
                session_timeout=90,
            ) as connection:
                connection.config_mode(pattern='(config)')
                connection.send_config_from_file(
                    config_file=cfg_file,
                    exit_config_mode=False,
                    enter_config_mode=False
                )

                output = connection.send_command('do sh run | i ip prefix-list BGP-OUT seq')
                re_result = list(filter(compiled_pattern_to_search.search, output.split('\n')))

                connection.send_config_set(f'no {re_result[0]}')

                connection.exit_config_mode(pattern='SITE')
                connection.save_config()

    except NetMikoTimeoutException:
        main_logger.info(f"Timeout exception: {equipment['secret']} ({equipment['ip']})")
        write_error(equipment)

    except AuthenticationException:
        main_logger.info(f"Authentication failed: {equipment['secret']} ({equipment['ip']})")
        write_error(equipment)

    except SSHException:
        main_logger.info(f"Error reading SSH protocol banner: {equipment['secret']} ({equipment['ip']})")
        write_error(equipment)

    except:
        main_logger.debug(f"Error: {equipment['secret']} ({equipment['ip']})")
        write_error(equipment)

    else:
        main_logger.info(f"Completed: {equipment['secret']} ({equipment['ip']})")
        write_output(equipment)


if __name__ == '__main__':
    cls()
    time_start = datetime.datetime.now()
    time_start = time_start.strftime('%H:%M:%S %d/%m/%Y')
    perf_counter_start = time.perf_counter()

    num_equips = main()

    perf_counter_stop = time.perf_counter()
    time_stop = datetime.datetime.now()
    time_stop = time_stop.strftime('%H:%M:%S %d/%m/%Y')

    main_logger.info(f'Number of equipments: {num_equips}')
    main_logger.info(f'Finished in {(perf_counter_stop - perf_counter_start):.3f} second(s)')
    main_logger.info(f'Started at {time_start} and finished at {time_stop}\n')
