# -*- coding: utf-8 -*-
import threading, re, xlrd, xlutils, xlwt, datetime, os, logging
from netmiko.ssh_exception import  NetMikoTimeoutException
from paramiko.ssh_exception import SSHException
from netmiko.ssh_exception import  AuthenticationException

logging.basicConfig(filename='../test.log', level=logging.DEBUG)
logger = logging.getLogger('netmiko')

from xlutils.copy import copy as xl_copy
from xlrd import open_workbook
from xlutils.save import save
from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import *

datatemp = str(datetime.today())
datatemp = re.search('(\S+)-(\S+)-(\S+)', datatemp)
data = datatemp.group(3) + '-' + datatemp.group(2) + '-' + datatemp.group(1)

wb = xlrd.open_workbook('C:\\Users\\xgftb21\\Documents\\GitHub\\netmiko-learning\\Rui Pereira\\Cadastro.xlsm') # Source Excel
wbbook = wb.sheet_by_name(u'Cadastro') # Sheet source excel

row = 1 # Numero da linha onde comecam os dados
coluna_testar = 16 # Numero da coluna Testar
linha = 1

path = 'D:\\rpereira\\Configs\\'

password_temp = open('D:\\rpereira\\pwd_tacacs.txt','r')
password = password_temp.read()

for i in range(1,wbbook.nrows):
    i+=1
ultima_coluna = i

def comandos_cpe(a_device):
    try:
        ssh = ConnectHandler(**a_device)
        host = ssh.find_prompt()
        linha_1 = (a_device)['secret']
        site_id = lista[linha_1-2]['site_id']
        cc = lista[linha_1 - 2]['cc']
        hostname_cpe = (a_device)['host']

        sh_dhcp_relay = ssh.send_command_expect('show run | i ip helper-address', host)
        for i in range(0,len(sh_dhcp_relay)):
            relay_temp = re.search('ip helper-address (\d+.\d+.\d+.\d+)', sh_dhcp_relay[i])
        print (host + ' ' + sh_dhcp_relay)

        ssh.disconnect()

    except Exception as falha:
        linha_1 = (a_device)['secret']
        site_id = lista[linha_1 - 2]['site_id']
        loopback = lista[linha_1 - 2]['ip']
        print(site_id + ' ' + loopback + ' ' + str(falha))

device_list = []
lista = []
linha_output_excel = 1
for row in range(1,ultima_coluna):
    if str(wbbook.cell(row, 18).value) == "ARS Norte" and str(wbbook.cell(row, 5).value) == "Migrado":
        site_id = str(wbbook.cell(row, 0).value)
        nome = str(wbbook.cell(row, 2).value)
        hostname = str(wbbook.cell(row, 7).value)
        loopback = str(wbbook.cell(row, 8).value)
        cc = str(wbbook.cell(row, 6).value)
        linha_output_excel +=1
        device_list.append({'host': hostname,
                            'ip': loopback ,
                            'username': 'rfpereira',
                            'password': password,
                            'secret' : linha_output_excel,
                            'alt_key_file' : cc,
                            'device_type': 'cisco_xr'})

        lista.append({'host': hostname,
                      'site_id' : site_id,
                      'ip': loopback,
                      'cc': cc,
                      'linha': linha_output_excel})

def main():
    n = 1
    for device in device_list:
        try:
            thread = threading.Thread(target=comandos_cpe, args=(device,))
            thread.start()
            if n==50:
                thread.join()
                n=1
            else:
                n+=1
        except:
            print ('Sem conectividade com o ' + str(device))
            n += 1

if __name__ == "__main__":
    main()
